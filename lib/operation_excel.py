#coding:utf-8

"""
封装操作excel类，使用openpyxl库代替xlrd
"""

import xlrd
import os
import configparser
from xlutils.copy import copy
from openpyxl import Workbook,load_workbook


class OperationExcel(object):
    def __init__(self, datadir, dirsec, filedir, namesec, filename, sheetID):
        self.conf = configparser.ConfigParser()
        self.conf.read(datadir)
        self.file = os.path.join(self.conf.get(dirsec, filedir), self.conf.get(namesec, filename))
        # self.tables = xlrd.open_workbook(self.file)
        self.tables = load_workbook(self.file)
        self.sheetID = sheetID
        self.table = self.tables[self.sheetID]

    def get_table_rows(self):
        """
        获取总行数
        :return:
        """
        return self.table.max_row

    def get_table_cols(self):
        """
        获取总列数
        :return:
        """
        return self.table.max_column

    def get_cell_value(self, row, col):
        """
        获取单元格的值
        :param row: 行号
        :param col: 列号
        :return:
        """
        return self.table.cell(row=int(row), column=int(col)).value

    def write_cell_value(self, row, col, value):
        """
        把数据写入单元格
        :param row:
        :param col:
        :param value:
        :return:
        """
        # tablescopy = copy(self.tables)
        # tablecopy = tablescopy.get_sheet(self.sheetID)
        # tablecopy.write(int(row), int(col), value)
        # tablescopy.save(self.file)
        self.table.cell(row=int(row), column=int(col), value=value)
        self.tables.save(self.file)

    def get_col_values(self, col):
        """
        获取某列所有值，除去第一行
        :param col:
        :return:
        """
        rows = self.get_table_rows()
        columndata = []
        for r in range(1, rows+1):
            cellvalue = self.get_cell_value(r, col)
            columndata.append(cellvalue)
        return columndata

    def get_row_values(self, row):
        """
        获取某行所有值
        :param row:
        :return:
        """
        columns = self.get_table_cols()
        rowdata = []
        for c in range(1,columns+1):
            cellvalue = self.get_cell_value(row, c)
            rowdata.append(cellvalue)
        return rowdata

